#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate a Google-Sheets-ready template for BehaveAI time-codes.

BehaveAI reads a CSV of frames to annotate (see parse_timecode_csv in
BehaveAI_annotation.py). This script writes an .xlsx template that you can
upload to / open in Google Sheets, fill in, then export back to .csv via
"File > Download > Comma-separated values (.csv)".

Why an .xlsx instead of a plain .csv:
  In Google Sheets a cell typed as "02:15" is auto-converted to a time value
  and exported as "02:15:00", which BehaveAI would read as 2h15m instead of
  2m15s. This template forces the timecode column to TEXT ("@") so values stay
  exactly as typed.

Usage:
  python make_timecodes_template.py [output.xlsx]

Default output: BehaveAI_timecodes_template.xlsx next to this script.
"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment
except ImportError:
    sys.exit(
        "openpyxl is required to build the template.\n"
        "Install it with:  pip install openpyxl"
    )

# Columns BehaveAI recognises (case-insensitive). The recommended header names
# are the first candidate of each group in parse_timecode_csv().
HEADERS = ["video_filename", "timecode", "behaviour"]

EXAMPLES = [
    ["my_video_01.mp4", "1530",     "grazing"],
    ["my_video_01.mp4", "02:15",    "walking"],
    ["my_video_02.mp4", "00:42",    "fighting"],
    ["my_video_02.mp4", "01:02:03", "resting"],
]

INSTRUCTIONS = [
    ("BehaveAI - Template de time-codes", True),
    ("", False),
    ("But : lister les images (frames) a annoter, une par ligne.", False),
    ("", False),
    ("Colonnes (onglet 'timecodes', ne pas renommer) :", True),
    ("  - video_filename : nom du fichier video dans clips/ (avec ou sans extension).", False),
    ("  - timecode       : numero de frame entier (ex. 1530)  OU  mm:ss (ex. 02:15)  OU  hh:mm:ss (ex. 01:02:03).", False),
    ("  - behaviour      : memo optionnel (ignore par l'outil).", False),
    ("", False),
    ("IMPORTANT - colonne timecode :", True),
    ("  La colonne B est formatee en TEXTE pour que Google Sheets ne transforme", False),
    ("  pas '02:15' en une heure. Gardez ce format texte si vous ajoutez des lignes.", False),
    ("  En cas de doute, utilisez des numeros de frame entiers : c'est sans ambiguite.", False),
    ("", False),
    ("Export pour BehaveAI :", True),
    ("  Fichier > Telecharger > Valeurs separees par des virgules (.csv)", False),
    ("  Seul l'onglet actif est exporte : placez-vous sur l'onglet 'timecodes' avant d'exporter.", False),
    ("  Placez le .csv dans le dossier projects/<projet>/timecodes/.", False),
    ("", False),
    ("Notes :", True),
    ("  - Les lignes vides et celles commencant par # sont ignorees.", False),
    ("  - Le nom de video est compare sans tenir compte de la casse ni de l'extension.", False),
    ("  - Les frames hors limites sont ramenees dans l'intervalle valide.", False),
    ("  - Les doublons (meme video + meme frame) sont supprimes.", False),
]


def build_template(out_path):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: timecodes (the data BehaveAI reads) ----
    ws = wb.active
    ws.title = "timecodes"
    ws.append(HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in EXAMPLES:
        ws.append(row)

    # Force the timecode column (B) to TEXT so Google Sheets keeps values as-is.
    for r in range(1, 2000):
        ws.cell(row=r, column=2).number_format = "@"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20

    for r in range(2, 2 + len(EXAMPLES)):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="left")

    ws["B1"].comment = Comment(
        "Format en TEXTE.\n"
        "Numero de frame entier (1530)\n"
        "OU mm:ss (02:15)\n"
        "OU hh:mm:ss (01:02:03).",
        "BehaveAI",
    )
    ws.freeze_panes = "A2"

    # ---- Sheet 2: instructions ----
    ins = wb.create_sheet("lisez-moi")
    ins.column_dimensions["A"].width = 100
    for i, (text, bold) in enumerate(INSTRUCTIONS, start=1):
        cell = ins.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)

    wb.save(out_path)
    return out_path


def main():
    if len(sys.argv) > 1:
        out_path = sys.argv[1]
    else:
        out_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "BehaveAI_timecodes_template.xlsx",
        )
    build_template(out_path)
    print(f"Template written to: {out_path}")


if __name__ == "__main__":
    main()
